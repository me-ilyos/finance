from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views import View
from django.views.generic import ListView
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Q, Count, Sum, Case, When, Value, DecimalField
from django.utils import timezone
from datetime import timedelta
from .forms import LoginForm, SalespersonForm
from .models import Salesperson
from apps.sales.models import Sale
from .services import DashboardService, DateFilterService
from apps.accounting.models import FinancialAccount
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import logging
from django.db import transaction

logger = logging.getLogger(__name__)

# Create your views here.

class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('core:dashboard')
        form = LoginForm()
        return render(request, 'login.html', {'form': form})

    def post(self, request):
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('core:dashboard')
            else:
                form.add_error(None, "Foydalanuvchi nomi yoki parol noto'g'ri.")
        return render(request, 'login.html', {'form': form})


class SalespersonListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Salesperson
    template_name = 'core/salesperson_list.html'
    context_object_name = 'salespeople'
    paginate_by = 20
    login_url = '/core/login/'

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        queryset = super().get_queryset().select_related('user').order_by('-created_at')
        
        # Store filter parameters for context
        self.filter_period = self.request.GET.get('filter_period')
        self.date_filter = self.request.GET.get('date_filter')
        self.start_date = self.request.GET.get('start_date')
        self.end_date = self.request.GET.get('end_date')

        # Apply date filtering using service for creation date
        start_date, end_date = DateFilterService.get_date_range(
            self.filter_period, self.date_filter, self.start_date, self.end_date
        )
        
        # Store date range for sales filtering
        self.sales_start_date = start_date
        self.sales_end_date = end_date
        
        # Annotate with sales statistics for the filtered period
        queryset = queryset.annotate(
            total_sales_count=Count(
                'sales_made',
                filter=Q(sales_made__sale_date__date__range=[start_date, end_date])
            ),
            total_sales_uzs=Sum(
                Case(
                    When(
                        sales_made__sale_date__date__range=[start_date, end_date],
                        sales_made__sale_currency='UZS',
                        then='sales_made__total_sale_amount'
                    ),
                    default=Value(0),
                    output_field=DecimalField()
                )
            ),
            total_sales_usd=Sum(
                Case(
                    When(
                        sales_made__sale_date__date__range=[start_date, end_date],
                        sales_made__sale_currency='USD',
                        then='sales_made__total_sale_amount'
                    ),
                    default=Value(0),
                    output_field=DecimalField()
                )
            )
        )
        
        # Don't filter by creation date for salespeople - show all salespeople with their sales stats
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if 'salesperson_form' not in context:
            context['salesperson_form'] = SalespersonForm()
        
        # Use the service to get filter context
        filter_context = DateFilterService.get_filter_context(
            self.filter_period, self.date_filter, self.start_date, self.end_date
        )
        context.update(filter_context)

        # Calculate statistics for salespeople
        filtered_queryset = self.get_queryset()
        
        # Calculate totals for the footer
        total_sales_uzs = sum(sp.total_sales_uzs or 0 for sp in filtered_queryset)
        total_sales_usd = sum(sp.total_sales_usd or 0 for sp in filtered_queryset)
        total_sales_count = sum(sp.total_sales_count or 0 for sp in filtered_queryset)
        
        stats = {
            'total_salespeople': filtered_queryset.count(),
            'active_salespeople': filtered_queryset.filter(is_active=True).count(),
            'inactive_salespeople': filtered_queryset.filter(is_active=False).count(),
            'total_sales_uzs': total_sales_uzs,
            'total_sales_usd': total_sales_usd,
            'total_sales_count': total_sales_count,
        }
        context['stats'] = stats
        
        # Add sales date range to context for display
        context['sales_date_range'] = f"{self.sales_start_date.strftime('%d.%m.%Y')} - {self.sales_end_date.strftime('%d.%m.%Y')}"
        
        # Preserve query parameters for pagination
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        return context

    def post(self, request, *args, **kwargs):
        """Handle salesperson creation"""
        form = SalespersonForm(request.POST)
        if form.is_valid():
            try:
                salesperson = form.save()
                if salesperson:
                    messages.success(request, "Yangi sotuvchi muvaffaqiyatli qo'shildi.")
                    return redirect(reverse_lazy('core:salesperson-list'))
                else:
                    messages.error(request, "Sotuvchini saqlashda xatolik yuz berdi.")
            except Exception as e:
                messages.error(request, f"Sotuvchini saqlashda xatolik: {e}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

        # Re-render with form errors
        self.object_list = self.get_queryset()
        context = self.get_context_data(salesperson_form=form, object_list=self.object_list)
        
        if not messages.get_messages(request):
            messages.error(request, "Sotuvchini qo'shishda xatolik. Iltimos, ma'lumotlarni tekshiring.")
        
        return self.render_to_response(context)


@login_required
def dashboard_view(request):
    """Dashboard view using service layer for data aggregation"""
    selected_account = None
    selected_account_id = request.GET.get('account_id')

    if selected_account_id:
        try:
            selected_account = get_object_or_404(
                FinancialAccount, 
                id=selected_account_id, 
                is_active=True
            )
        except FinancialAccount.DoesNotExist:
            messages.error(request, "Tanlangan hisob topilmadi.")
            return redirect('core:dashboard')

    # Get pagination parameters
    transactions_page = request.GET.get('page', 1)
    transactions_per_page = 10

    # Get dashboard data using service
    dashboard_data = DashboardService.get_dashboard_data(
        selected_account, 
        transactions_page, 
        transactions_per_page
    )
    
    context = {
        **dashboard_data,
        'page_title': 'Boshqaruv Paneli'
    }
    
    return render(request, 'core/dashboard.html', context)

@login_required
def logout_view(request):
    logout(request)
    return redirect('core:login')

@login_required
def salesperson_export_excel(request):
    """Export salesperson data to Excel (XLSX) format"""
    if not request.user.is_superuser:
        messages.error(request, "Bu funktsiyaga faqat administratorlar kirish huquqiga ega.")
        return redirect('core:salesperson-list')
    
    try:
        # Get the same queryset as the list view
        queryset = Salesperson.objects.select_related('user').order_by('-created_at')
        
        # Apply the same filtering as the list view
        filter_period = request.GET.get('filter_period')
        date_filter = request.GET.get('date_filter')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        start_date_obj, end_date_obj = DateFilterService.get_date_range(
            filter_period, date_filter, start_date, end_date
        )
        
        # Annotate with sales statistics for the filtered period
        queryset = queryset.annotate(
            total_sales_count=Count(
                'sales_made',
                filter=Q(sales_made__sale_date__date__range=[start_date_obj, end_date_obj])
            ),
            total_sales_uzs=Sum(
                Case(
                    When(
                        sales_made__sale_date__date__range=[start_date_obj, end_date_obj],
                        sales_made__sale_currency='UZS',
                        then='sales_made__total_sale_amount'
                    ),
                    default=Value(0),
                    output_field=DecimalField()
                )
            ),
            total_sales_usd=Sum(
                Case(
                    When(
                        sales_made__sale_date__date__range=[start_date_obj, end_date_obj],
                        sales_made__sale_currency='USD',
                        then='sales_made__total_sale_amount'
                    ),
                    default=Value(0),
                    output_field=DecimalField()
                )
            )
        )

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sotuvchilar Hisoboti"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "№", 
            "To'liq Ism", 
            "Foydalanuvchi Nomi", 
            "Telefon", 
            "Sotuvlar Soni", 
            "Jami Sotuv (UZS)", 
            "Jami Sotuv (USD)",
            "Holat",
            "Yaratilgan Sana"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_idx, salesperson in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Row number
            ws.cell(row=row_idx, column=2, value=salesperson.user.get_full_name() or "N/A")
            ws.cell(row=row_idx, column=3, value=salesperson.user.username)
            ws.cell(row=row_idx, column=4, value=salesperson.phone_number or "N/A")
            ws.cell(row=row_idx, column=5, value=salesperson.total_sales_count or 0)
            ws.cell(row=row_idx, column=6, value=float(salesperson.total_sales_uzs or 0))
            ws.cell(row=row_idx, column=7, value=float(salesperson.total_sales_usd or 0))
            ws.cell(row=row_idx, column=8, value="Faol" if salesperson.is_active else "Faol emas")
            ws.cell(row=row_idx, column=9, value=salesperson.created_at.strftime('%d.%m.%Y %H:%M') if salesperson.created_at else "N/A")

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(
                len(str(ws.cell(row=1, column=col).value)),
                max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(2, ws.max_row + 1)) if ws.max_row > 1 else 0
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Add filter information
        filter_info = f"Filtrlash davri: {start_date_obj.strftime('%d.%m.%Y')} - {end_date_obj.strftime('%d.%m.%Y')}"
        ws.cell(row=ws.max_row + 2, column=1, value=filter_info)

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sotuvchilar_hisoboti_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error exporting salespeople to Excel: {e}")
        messages.error(request, "Excel faylini eksport qilishda xatolik yuz berdi.")
        return redirect('core:salesperson-list')

@login_required
def salesperson_edit(request, salesperson_id):
    """Edit salesperson details"""
    if not request.user.is_superuser:
        messages.error(request, "Bu funktsiyaga faqat administratorlar kirish huquqiga ega.")
        return redirect('core:salesperson-list')
    
    try:
        salesperson = get_object_or_404(Salesperson, id=salesperson_id)
        
        if request.method == 'POST':
            # Get form data
            username = request.POST.get('username', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            phone_number = request.POST.get('phone_number', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            password = request.POST.get('password', '').strip()
            password_confirm = request.POST.get('password_confirm', '').strip()
            
            # Validate required fields
            if not username or not first_name or not last_name:
                messages.error(request, "Majburiy maydonlarni to'ldiring.")
                return redirect('core:salesperson-list')
            
            # Validate password if provided
            if password or password_confirm:
                if not password:
                    messages.error(request, "Parol maydonini to'ldiring.")
                    return redirect('core:salesperson-list')
                if password != password_confirm:
                    messages.error(request, "Parollar mos kelmadi.")
                    return redirect('core:salesperson-list')
                if len(password) < 8:
                    messages.error(request, "Parol kamida 8 ta belgidan iborat bo'lishi kerak.")
                    return redirect('core:salesperson-list')
            
            # Check if username is unique (excluding current user)
            from django.contrib.auth.models import User
            if User.objects.filter(username=username).exclude(id=salesperson.user.id).exists():
                messages.error(request, "Bu foydalanuvchi nomi allaqachon mavjud.")
                return redirect('core:salesperson-list')
            
            try:
                with transaction.atomic():
                    # Update user fields
                    user = salesperson.user
                    user.username = username
                    user.first_name = first_name
                    user.last_name = last_name
                    user.email = email
                    
                    # Update password if provided
                    if password:
                        user.set_password(password)
                    
                    user.save()
                    
                    # Update salesperson fields
                    salesperson.phone_number = phone_number
                    salesperson.is_active = is_active
                    salesperson.save()
                    
                    success_message = "Sotuvchi ma'lumotlari muvaffaqiyatli yangilandi."
                    if password:
                        success_message += " Parol ham o'zgartirildi."
                    messages.success(request, success_message)
                    
            except Exception as e:
                logger.error(f"Error updating salesperson {salesperson_id}: {e}")
                messages.error(request, "Yangilashda xatolik yuz berdi.")
        
        return redirect('core:salesperson-list')
        
    except Exception as e:
        logger.error(f"Error in salesperson_edit: {e}")
        messages.error(request, "Sotuvchini tahrirlashda xatolik yuz berdi.")
        return redirect('core:salesperson-list')


@login_required  
def salesperson_toggle_status(request, salesperson_id):
    """Toggle salesperson active status"""
    if not request.user.is_superuser:
        messages.error(request, "Bu funktsiyaga faqat administratorlar kirish huquqiga ega.")
        return redirect('core:salesperson-list')
    
    if request.method == 'POST':
        try:
            salesperson = get_object_or_404(Salesperson, id=salesperson_id)
            action = request.POST.get('action')
            
            if action == 'activate':
                salesperson.is_active = True
                action_text = "faollashtirildi"
            elif action == 'deactivate':
                salesperson.is_active = False
                action_text = "faolsizlashtirildi"
            else:
                messages.error(request, "Noto'g'ri amal.")
                return redirect('core:salesperson-list')
            
            salesperson.save(update_fields=['is_active', 'updated_at'])
            
            full_name = f"{salesperson.user.first_name} {salesperson.user.last_name}".strip()
            if not full_name:
                full_name = salesperson.user.username
                
            messages.success(request, f"Sotuvchi {full_name} muvaffaqiyatli {action_text}.")
            
        except Exception as e:
            logger.error(f"Error toggling salesperson status {salesperson_id}: {e}")
            messages.error(request, "Holat o'zgartirishda xatolik yuz berdi.")
    
    return redirect('core:salesperson-list')

class SalespersonDetailView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Sale
    template_name = 'core/salesperson_detail.html'
    context_object_name = 'sales'
    paginate_by = 20
    login_url = '/core/login/'

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        self.salesperson = get_object_or_404(Salesperson, id=self.kwargs['salesperson_id'])
        
        queryset = Sale.objects.filter(salesperson=self.salesperson).select_related(
            'related_acquisition', 
            'related_acquisition__ticket',
            'agent', 
            'paid_to_account',
            'salesperson',
            'salesperson__user'
        ).order_by('-sale_date', '-created_at')
        
        # Store filter parameters for context
        self.filter_period = self.request.GET.get('filter_period')
        self.date_filter = self.request.GET.get('date_filter')
        self.start_date = self.request.GET.get('start_date')
        self.end_date = self.request.GET.get('end_date')

        # Apply date filtering using service
        start_date, end_date = DateFilterService.get_date_range(
            self.filter_period, self.date_filter, self.start_date, self.end_date
        )
        
        return queryset.filter(sale_date__date__range=[start_date, end_date])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['salesperson'] = self.salesperson
        
        # Use the service to get filter context
        filter_context = DateFilterService.get_filter_context(
            self.filter_period, self.date_filter, self.start_date, self.end_date
        )
        context.update(filter_context)

        # Calculate totals for the current filtered queryset
        filtered_queryset = self.get_queryset()
        totals = filtered_queryset.aggregate(
            total_quantity=Sum('quantity'),
            total_sum_uzs=Sum(
                Case(When(sale_currency='UZS', then='total_sale_amount'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_sum_usd=Sum(
                Case(When(sale_currency='USD', then='total_sale_amount'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_profit_uzs=Sum(
                Case(When(sale_currency='UZS', then='profit'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_profit_usd=Sum(
                Case(When(sale_currency='USD', then='profit'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_initial_payment_uzs=Sum(
                Case(When(agent__isnull=False, sale_currency='UZS', then='initial_payment_amount'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_initial_payment_usd=Sum(
                Case(When(agent__isnull=False, sale_currency='USD', then='initial_payment_amount'), 
                     default=Value(0), output_field=DecimalField())
            )
        )
        context['totals'] = totals
        
        # Preserve query parameters for pagination
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        return context
