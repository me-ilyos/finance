from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone


class ExcelExportService:
    @staticmethod
    def export_salespeople(queryset, start_date, end_date):
        """Export salesperson data to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sotuvchilar Hisoboti"

        # Headers
        headers = [
            "№", "To'liq Ism", "Foydalanuvchi Nomi", "Telefon", 
            "Sotuvlar Soni", "Jami Sotuv (UZS)", "Jami Sotuv (USD)",
            "Holat", "Yaratilgan Sana"
        ]
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, salesperson in enumerate(queryset, 2):
            row_data = [
                row_idx - 1,  # Row number
                salesperson.user.get_full_name() or "N/A",
                salesperson.user.username,
                salesperson.phone_number or "N/A",
                salesperson.total_sales_count or 0,
                float(salesperson.total_sales_uzs or 0),
                float(salesperson.total_sales_usd or 0),
                "Faol" if salesperson.is_active else "Faol emas",
                salesperson.created_at.strftime('%d.%m.%Y %H:%M') if salesperson.created_at else "N/A"
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(
                len(str(ws.cell(row=1, column=col).value)),
                max(len(str(ws.cell(row=row, column=col).value or "")) for row in range(2, ws.max_row + 1)) if ws.max_row > 1 else 0
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Add filter information
        filter_info = f"Filtrlash davri: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws.cell(row=ws.max_row + 2, column=1, value=filter_info)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sotuvchilar_hisoboti_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        
        wb.save(response)
        return response 