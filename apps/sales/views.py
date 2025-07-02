from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from .models import Sale
from .forms import SaleForm
from apps.inventory.models import Acquisition
from apps.accounting.models import FinancialAccount
from apps.core.models import Salesperson
from apps.contacts.models import AgentPayment
from django.db.models import Sum, Case, When, Value, DecimalField
from django.core.exceptions import ValidationError
from apps.core.services import DateFilterService
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)


class SaleListView(LoginRequiredMixin, ListView):
    model = Sale
    template_name = 'sales/sale_list.html'
    context_object_name = 'sales'
    paginate_by = 20
    login_url = '/core/login/'

    def get_queryset(self):
        return self._get_filtered_queryset()

    def _get_filtered_queryset(self):
        """Centralized queryset filtering to avoid duplication"""
        queryset = self.model.objects.select_related(
            'related_acquisition', 
            'related_acquisition__ticket',
            'agent', 
            'paid_to_account',
            'salesperson',
            'salesperson__user'
        ).order_by('-sale_date', '-created_at')
        
        # Filter by current salesperson - only show their own sales
        try:
            current_salesperson = self.request.user.salesperson_profile
            queryset = queryset.filter(salesperson=current_salesperson)
        except Salesperson.DoesNotExist:
            # If user is not a salesperson (admin/superuser), show all sales
            if not self.request.user.is_superuser:
                # Non-salesperson, non-admin users see no sales
                queryset = queryset.none()
        
        # Apply date filtering using service
        try:
            start_date, end_date = DateFilterService.get_date_range(
                self.request.GET.get('filter_period'),
                self.request.GET.get('date_filter'),
                self.request.GET.get('start_date'),
                self.request.GET.get('end_date')
            )
            queryset = queryset.filter(sale_date__date__range=[start_date, end_date])
        except ValueError as e:
            logger.warning(f"Date filter error: {e}")
            today = timezone.localdate()
            queryset = queryset.filter(sale_date__date=today)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if 'sale_form' not in context:
            context['sale_form'] = SaleForm(current_user=self.request.user)
        
        # Use the service to get filter context
        filter_context = DateFilterService.get_filter_context(
            self.request.GET.get('filter_period'),
            self.request.GET.get('date_filter'),
            self.request.GET.get('start_date'),
            self.request.GET.get('end_date')
        )
        context.update(filter_context)

        # Calculate totals for the current filtered queryset
        filtered_queryset = self.get_queryset()
        totals = filtered_queryset.aggregate(
            total_quantity=Sum('quantity'),
            total_sum_uzs=Sum(
                Case(When(sale_currency='UZS', then='total_sale_amount'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_sum_usd=Sum(
                Case(When(sale_currency='USD', then='total_sale_amount'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_profit_uzs=Sum(
                Case(When(sale_currency='UZS', then='profit'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_profit_usd=Sum(
                Case(When(sale_currency='USD', then='profit'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_initial_payment_uzs=Sum(
                Case(When(agent__isnull=False, sale_currency='UZS', then='initial_payment_amount'), 
                     default=Value(0), output_field=DecimalField())
            ),
            total_initial_payment_usd=Sum(
                Case(When(agent__isnull=False, sale_currency='USD', then='initial_payment_amount'), 
                     default=Value(0), output_field=DecimalField())
            )
        )
        context['totals'] = totals
        
        # Preserve query parameters for pagination
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        return context

    def post(self, request, *args, **kwargs):
        """Handle sale creation with business logic"""
        form = SaleForm(request.POST, current_user=request.user)
        if form.is_valid():
            try:
                sale = self._create_sale(form)
                if sale:
                    messages.success(request, "Yangi sotuv muvaffaqiyatli qo'shildi.")
                    return redirect(reverse_lazy('sales:sale-list'))
                else:
                    messages.error(request, "Sotuvni saqlashda xatolik yuz berdi.")
            except ValidationError as e:
                logger.error(f"Validation error creating sale: {e}")
                messages.error(request, f"Sotuvni saqlashda xatolik: {e}")
            except Exception as e:
                logger.error(f"Unexpected error creating sale: {e}")
                messages.error(request, "Kutilmagan xatolik yuz berdi.")
        else:
            logger.warning(f"Invalid form data: {form.errors}")

        # Re-render with form errors
        self.object_list = self.get_queryset()
        context = self.get_context_data(sale_form=form, object_list=self.object_list)
        
        if not messages.get_messages(request):
            messages.error(request, "Sotuvni qo'shishda xatolik. Iltimos, ma'lumotlarni tekshiring.")
        
        return self.render_to_response(context)

    def _create_sale(self, form):
        """Create sale with business logic (moved from services)"""
        with transaction.atomic():
            # Create sale instance
            sale = form.save(commit=False)
            
            # Set calculated fields
            sale.sale_currency = sale.related_acquisition.currency
            sale.total_sale_amount = sale.quantity * sale.unit_sale_price
            
            # Calculate profit
            unit_cost = sale.related_acquisition.unit_price
            total_cost = sale.quantity * unit_cost
            sale.profit = sale.total_sale_amount - total_cost
            
            # Add salesperson
            if 'salesperson' in form.cleaned_data:
                sale.salesperson = form.cleaned_data['salesperson']
            
            sale.save()
            
            # Update stock
            acquisition = sale.related_acquisition
            acquisition.available_quantity -= sale.quantity
            acquisition.save(update_fields=['available_quantity', 'updated_at'])
            
            # Handle agent debt
            if sale.agent:
                sale.agent.add_debt(sale.total_sale_amount, sale.sale_currency)
                
                # Create initial payment if provided
                if sale.initial_payment_amount and sale.initial_payment_amount > 0 and sale.paid_to_account:
                    AgentPayment.objects.create(
                        agent=sale.agent,
                        payment_date=sale.sale_date,
                        amount=sale.initial_payment_amount,
                        currency=sale.sale_currency,
                        paid_to_account=sale.paid_to_account,
                        notes=f"Avtomatik boshlang'ich to'lov (Sotuv ID: {sale.id})"
                    )
            
            # Handle client payment
            elif sale.paid_to_account:
                account = sale.paid_to_account
                account.current_balance += sale.total_sale_amount
                account.save(update_fields=['current_balance', 'updated_at'])
            
            return sale


@login_required(login_url='/core/login/')
def get_accounts_for_acquisition_currency(request, acquisition_id):
    """AJAX view to get accounts matching acquisition currency"""
    try:
        acquisition = Acquisition.objects.get(pk=acquisition_id)
        currency = acquisition.currency
        accounts = FinancialAccount.objects.filter(
            currency=currency, 
            is_active=True
        ).values('id', 'name')
        
        return JsonResponse(list(accounts), safe=False)
        
    except Acquisition.DoesNotExist:
        return JsonResponse({'error': 'Acquisition not found'}, status=404)
    except Exception as e:
        logger.error(f"Error in get_accounts_for_acquisition_currency: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/core/login/')
def export_sales_excel(request):
    """Export sales data to Excel format"""
    try:
        # Use the same filtering logic as the list view
        queryset = Sale.objects.select_related(
            'related_acquisition', 
            'related_acquisition__ticket',
            'agent', 
            'paid_to_account',
            'salesperson',
            'salesperson__user'
        ).order_by('-sale_date', '-created_at')
        
        # Filter by current salesperson
        try:
            current_salesperson = request.user.salesperson_profile
            queryset = queryset.filter(salesperson=current_salesperson)
        except Salesperson.DoesNotExist:
            if not request.user.is_superuser:
                queryset = queryset.none()
        
        # Apply date filtering
        try:
            start_date_obj, end_date_obj = DateFilterService.get_date_range(
                request.GET.get('filter_period'),
                request.GET.get('date_filter'),
                request.GET.get('start_date'),
                request.GET.get('end_date')
            )
            queryset = queryset.filter(sale_date__date__range=[start_date_obj, end_date_obj])
        except ValueError:
            today = timezone.localdate()
            queryset = queryset.filter(sale_date__date=today)
            start_date_obj = end_date_obj = today

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sotuvlar Ro'yxati"

        # Headers
        headers = [
            'Sana', 'Chipta turi', 'Chipta manzili', 'Xaridor',
            'Miqdori', 'Birlik narxi', 'Jami summa', 'Valyuta',
            'Foyda', 'Boshlang\'ich to\'lov', 'To\'lov hisobi', 'To\'lov holati', 'Agent'
        ]

        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_num, sale in enumerate(queryset, 2):
            # Format data
            buyer = sale.agent.name if sale.agent else (sale.client_full_name or "N/A")
            agent_name = sale.agent.name if sale.agent else ""
            
            # Format prices based on currency
            if sale.sale_currency == 'UZS':
                unit_price = f"{sale.unit_sale_price:,.0f}"
                total_amount = f"{sale.total_sale_amount:,.0f}"
                profit = f"{sale.profit:,.0f}"
                initial_payment = f"{sale.initial_payment_amount:,.0f}" if sale.initial_payment_amount else ""
            else:  # USD
                unit_price = f"{sale.unit_sale_price:,.2f}"
                total_amount = f"{sale.total_sale_amount:,.2f}"
                profit = f"{sale.profit:,.2f}"
                initial_payment = f"{sale.initial_payment_amount:,.2f}" if sale.initial_payment_amount else ""
            
            # Payment status
            if sale.agent:
                payment_status = "Agent qarzi"
                payment_account = ""
            elif sale.paid_to_account:
                payment_status = "To'langan"
                payment_account = sale.paid_to_account.name
            else:
                payment_status = "To'lanmagan"
                payment_account = ""

            row_data = [
                sale.sale_date.strftime('%d.%m.%Y %H:%M'),
                sale.related_acquisition.ticket.get_ticket_type_display(),
                sale.related_acquisition.ticket.description,
                buyer,
                sale.quantity,
                unit_price,
                total_amount,
                sale.sale_currency,
                profit,
                initial_payment,
                payment_account,
                payment_status,
                agent_name
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Center align numeric and status columns
                if col_num in [5, 6, 7, 8, 9, 10, 11, 12]:
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Set filename
        if start_date_obj == end_date_obj:
            filename = f"sotuvlar_{start_date_obj.strftime('%d.%m.%Y')}.xlsx"
        else:
            filename = f"sotuvlar_{start_date_obj.strftime('%d.%m.%Y')}_dan_{end_date_obj.strftime('%d.%m.%Y')}_gacha.xlsx"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        logger.info(f"Exported {queryset.count()} sales to Excel")
        return response

    except Exception as e:
        logger.error(f"Error exporting sales to Excel: {e}")
        return HttpResponse("Excel export xatolik bilan yakunlandi.", status=500)
