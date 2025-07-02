from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import ListView
from django.http import HttpResponse
from django.contrib import messages
from .models import Acquisition, Ticket
from .forms import AcquisitionForm
from apps.core.services import DateFilterService
from apps.contacts.models import SupplierPayment
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class AcquisitionListView(ListView):
    model = Acquisition
    template_name = 'inventory/acquisition_list.html'
    context_object_name = 'acquisitions'
    paginate_by = 20

    def get_queryset(self):
        return self._get_filtered_queryset()

    def _get_filtered_queryset(self):
        """Centralized queryset filtering to avoid duplication"""
        queryset = self.model.objects.select_related(
            'supplier', 'ticket', 'paid_from_account'
        ).order_by('-acquisition_date', '-created_at')
        
        # Apply date filtering
        try:
            start_date_obj, end_date_obj = DateFilterService.get_date_range(
                self.request.GET.get('filter_period'),
                self.request.GET.get('date_filter'),
                self.request.GET.get('start_date'),
                self.request.GET.get('end_date')
            )
            queryset = queryset.filter(acquisition_date__date__range=[start_date_obj, end_date_obj])
        except ValueError as e:
            logger.warning(f"Date filter error: {e}")
            # Fall back to today's data on error
            today = timezone.localdate()
            queryset = queryset.filter(acquisition_date__date=today)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add form to context
        if 'acquisition_form' not in context:
            context['acquisition_form'] = AcquisitionForm()
        
        # Use centralized filter context service
        filter_context = DateFilterService.get_filter_context(
            self.request.GET.get('filter_period'),
            self.request.GET.get('date_filter'),
            self.request.GET.get('start_date'),
            self.request.GET.get('end_date')
        )
        context.update(filter_context)
        
        # For pagination - preserve query parameters
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        return context

    def post(self, request, *args, **kwargs):
        """Handle acquisition creation with ticket and business logic"""
        form = AcquisitionForm(request.POST)
        if form.is_valid():
            try:
                # Create ticket first
                ticket = Ticket.objects.create(
                    ticket_type=form.cleaned_data['ticket_type'],
                    description=form.cleaned_data['ticket_description'],
                    departure_date_time=form.cleaned_data['ticket_departure_date_time'],
                    arrival_date_time=form.cleaned_data.get('ticket_arrival_date_time')
                )
                
                # Create acquisition
                acquisition = form.save(commit=False)
                acquisition.ticket = ticket
                acquisition.save()
                
                # Handle business logic moved from model
                # Step 1: Always add debt to supplier
                acquisition.supplier.add_debt(acquisition.total_amount, acquisition.currency)
                
                # Step 2: If automatic payment, create payment record
                if acquisition.paid_from_account:
                    SupplierPayment.objects.create(
                        supplier=acquisition.supplier,
                        payment_date=acquisition.acquisition_date,
                        amount=acquisition.total_amount,
                        currency=acquisition.currency,
                        paid_from_account=acquisition.paid_from_account,
                        notes=f"Avtomatik to'lov - Xarid #{acquisition.pk}"
                    )
                
                messages.success(request, "Xarid muvaffaqiyatli qo'shildi.")
                logger.info(f"Created acquisition {acquisition.id}")
                return redirect(reverse_lazy('inventory:acquisition-list'))
            except Exception as e:
                logger.error(f"Error creating acquisition: {e}")
                messages.error(request, "Xarid yaratishda xatolik yuz berdi.")
        else:
            logger.warning(f"Form validation errors: {form.errors}")
            messages.error(request, "Ma'lumotlarni tekshiring.")
        
        # Re-render with form errors
        self.object_list = self.get_queryset()
        context = self.get_context_data(acquisition_form=form, object_list=self.object_list)
        return self.render_to_response(context)


def export_acquisitions_excel(request):
    """Export acquisitions data to Excel format"""
    try:
        # Use the same filtering logic as the list view
        queryset = Acquisition.objects.select_related(
            'supplier', 'ticket', 'paid_from_account'
        ).order_by('-acquisition_date', '-created_at')
        
        # Apply date filtering (reuse logic)
        try:
            start_date_obj, end_date_obj = DateFilterService.get_date_range(
                request.GET.get('filter_period'),
                request.GET.get('date_filter'),
                request.GET.get('start_date'),
                request.GET.get('end_date')
            )
            queryset = queryset.filter(acquisition_date__date__range=[start_date_obj, end_date_obj])
        except ValueError:
            today = timezone.localdate()
            queryset = queryset.filter(acquisition_date__date=today)
            start_date_obj = end_date_obj = today

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Xaridlar Ro'yxati"

        # Headers
        headers = [
            'Sana', 'Ta\'minotchi', 'Chipta turi', 'Chipta manzili',
            'Mavjud miqdori', 'Boshlang\'ich miqdori', 'Birlik narxi',
            'Jami summa', 'Valyuta', 'To\'lov hisobi', 'To\'lov holati', 'Izohlar'
        ]

        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_num, acq in enumerate(queryset, 2):
            # Format data
            row_data = [
                acq.acquisition_date.strftime('%d.%m.%Y'),
                acq.supplier.name,
                acq.ticket.get_ticket_type_display(),
                acq.ticket.description,
                acq.available_quantity,
                acq.initial_quantity,
                f"{acq.unit_price:,.0f}" if acq.currency == 'UZS' else f"{acq.unit_price:,.2f}",
                f"{acq.total_amount:,.0f}" if acq.currency == 'UZS' else f"{acq.total_amount:,.2f}",
                acq.currency,
                acq.paid_from_account.name if acq.paid_from_account else "To'lanmagan",
                "To'langan" if acq.paid_from_account else "To'lanmagan",
                acq.notes or ""
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Center align numeric and status columns
                if col_num in [5, 6, 7, 8, 9, 10, 11]:
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Set filename
        if start_date_obj == end_date_obj:
            filename = f"xaridlar_{start_date_obj.strftime('%d.%m.%Y')}.xlsx"
        else:
            filename = f"xaridlar_{start_date_obj.strftime('%d.%m.%Y')}_dan_{end_date_obj.strftime('%d.%m.%Y')}_gacha.xlsx"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        logger.info(f"Exported {queryset.count()} acquisitions to Excel")
        return response

    except Exception as e:
        logger.error(f"Error exporting acquisitions to Excel: {e}")
        return HttpResponse("Excel export xatolik bilan yakunlandi.", status=500)